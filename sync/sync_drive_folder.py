#!/usr/bin/env python3
"""
Sync Google Drive folder → data.json for Health Check Dashboard.

Scans a public Google Drive folder for .xlsx/.csv files,
downloads the most recent one, parses WebManagement export format,
and updates data.json.

Supports two WebManagement export formats:
  1. Grouped by country: rows = country × month (Country | Month | metrics)
  2. Flat by country: rows = country (Country | metrics)

Usage:
  python sync/sync_drive_folder.py

Env vars:
  DRIVE_FOLDER_ID — Google Drive folder ID
  DATA_JSON_PATH  — (optional) path to data.json
"""

import csv
import io
import json
import os
import re
import sys
import tempfile
import urllib.request
from datetime import datetime, timezone

DATA_JSON_PATH = os.environ.get("DATA_JSON_PATH", "health-check-dashboard/data.json")
DRIVE_FOLDER_ID = os.environ.get("DRIVE_FOLDER_ID", "1cyQ59HtuH3v2J5zB8DUbiOgXKqZyDazD")

# ─── GEO resolution ───

GEO_ALIASES = {
    "brazil": "🇧🇷 Brazil", "бразилия": "🇧🇷 Brazil", "brasil": "🇧🇷 Brazil", "br": "🇧🇷 Brazil",
    "argentina": "🇦🇷 Argentina", "аргентина": "🇦🇷 Argentina", "ar": "🇦🇷 Argentina",
    "chile": "🇨🇱 Chile", "чили": "🇨🇱 Chile", "cl": "🇨🇱 Chile",
    "mexico": "🇲🇽 Mexico", "мексика": "🇲🇽 Mexico", "mx": "🇲🇽 Mexico", "méxico": "🇲🇽 Mexico",
    "peru": "🇵🇪 Peru", "перу": "🇵🇪 Peru", "pe": "🇵🇪 Peru", "perú": "🇵🇪 Peru",
    "colombia": "🇨🇴 Colombia", "колумбия": "🇨🇴 Colombia", "co": "🇨🇴 Colombia",
    "india": "🇮🇳 India", "индия": "🇮🇳 India", "in": "🇮🇳 India",
    "bangladesh": "🇧🇩 Bangladesh", "бангладеш": "🇧🇩 Bangladesh", "bd": "🇧🇩 Bangladesh",
    "poland": "🇵🇱 Poland", "польша": "🇵🇱 Poland", "pl": "🇵🇱 Poland", "polska": "🇵🇱 Poland",
    "portugal": "🇵🇹 Portugal", "португалия": "🇵🇹 Portugal", "pt": "🇵🇹 Portugal",
}
KNOWN_GEOS = list({v for v in GEO_ALIASES.values()})


def resolve_geo(raw: str) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("total", "итого", "всего", ""):
        return None
    low = s.lower()
    if low in GEO_ALIASES:
        return GEO_ALIASES[low]
    for geo in KNOWN_GEOS:
        country = geo.split(" ", 1)[-1].lower()
        if country in low or low in country:
            return geo
    return None


# ─── WebManagement column mapping ───

WM_COLUMN_MAP = {
    "deposits amount": "deposits_total",
    "withdrawals amount": "withdrawals_total",
    "in-out": "in_out",
    "ggr": "ggr",
    "%in-out": "inout_pct",
    "%rtp casino": "rtp_casino",
    "registration": "registrations",
    "ftds": "ftd_count",
    "reg2dep, %": "reg_ftd_conv",
    "reg2dep,%": "reg_ftd_conv",
    "cr 2 dep, %": "cr2dep",
    "cr 3 dep, %": "cr3dep",
    "cr 4 dep, %": "cr4dep",
    "cr 5+ dep, %": "cr5dep",
}

DIRECT_METRIC_MAP = {
    "au": "au", "active unique": "au",
    "dep approval": "dep_approval", "deposit approval": "dep_approval",
    "wd approval": "wd_approval", "withdraw approval": "wd_approval",
    "avg deposit": "avg_deposit", "average deposit": "avg_deposit",
    "cashout": "cashout_pct", "cashout %": "cashout_pct",
    "ngr": "ngr", "net gaming revenue": "ngr",
    "ngr mom": "ngr_mom",
    "ngr target": "ngr_target", "ngr vs 150k": "ngr_target",
    "arppu": "arppu",
    "casino share": "casino_share",
    "casino margin": "casino_margin",
    "casino players": "casino_players",
    "casino ggr mom": "casino_ggr_mom",
    "sport margin": "sport_margin",
    "sport players": "sport_players",
    "sport ggr mom": "sport_ggr_mom",
    "avg bet": "avg_bet", "average bet": "avg_bet",
    "d30 retention": "d30_retention", "m1 ret": "d30_retention",
    "d90 retention": "d90_retention", "m3 ret": "d90_retention",
    "core retention": "core_retention",
    "active mom": "active_mom",
    "bettor pct": "bettor_pct",
    "active daily": "active_daily", "dau": "active_daily",
    "bettor ret": "bettor_ret",
    "whale top20": "whale_top20",
    "frt": "frt_sec",
    "csat": "csat",
    "abandon rate": "abandon_rate", "ar%": "abandon_rate",
    "aht": "aht_min",
}


def map_column(header: str) -> str:
    low = header.strip().lower()
    if low in WM_COLUMN_MAP:
        return WM_COLUMN_MAP[low]
    if low in DIRECT_METRIC_MAP:
        return DIRECT_METRIC_MAP[low]
    for key, metric in DIRECT_METRIC_MAP.items():
        if key in low:
            return metric
    return ""


def parse_number(raw) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip().replace("\xa0", "").replace(" ", "")
    if not s or s in ("-", "—", "–", "null", "n/a", "na", "None"):
        return None
    s = s.replace(",", ".")
    s = re.sub(r"[%$€]", "", s)
    try:
        return float(s)
    except ValueError:
        return None


# ─── Drive folder scanner ───

def list_drive_files(folder_id: str) -> list[dict]:
    """List files in a public Google Drive folder via embedded viewer."""
    url = f"https://drive.google.com/embeddedfolderview?id={folder_id}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=20) as resp:
        html = resp.read().decode("utf-8", errors="replace")

    files = []
    titles = re.findall(r'flip-entry-title">([^<]+)<', html)
    file_ids = re.findall(r'/file/d/([^/"]+)', html)

    for title, fid in zip(titles, file_ids):
        files.append({"name": title.strip(), "id": fid})

    print(f"[sync] Found {len(files)} files in Drive folder")
    for f in files:
        print(f"[sync]   {f['name']} ({f['id'][:12]}...)")
    return files


def download_drive_file(file_id: str, dest: str):
    """Download a file from Google Drive by ID."""
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = resp.read()
    # Handle virus scan warning for large files
    if b"confirm=" in data and b"download_warning" in data:
        confirm = re.search(rb"confirm=([0-9A-Za-z_-]+)", data)
        if confirm:
            url2 = f"{url}&confirm={confirm.group(1).decode()}"
            req2 = urllib.request.Request(url2, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req2, timeout=60) as resp2:
                data = resp2.read()
    with open(dest, "wb") as f:
        f.write(data)
    print(f"[sync] Downloaded {len(data)} bytes → {dest}")


def pick_best_file(files: list[dict]) -> dict | None:
    """Pick the most relevant file (xlsx/csv, prefer newest or with country data)."""
    candidates = [f for f in files if f["name"].lower().endswith((".xlsx", ".csv", ".tsv"))]
    if not candidates:
        return None
    candidates.sort(key=lambda f: f["name"], reverse=True)
    return candidates[0]


# ─── XLSX parser ───

def parse_xlsx(filepath: str) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(c or "").strip() for c in rows[0]]
        print(f"[sync] Sheet '{sheet_name}': {len(rows)-1} rows, headers: {headers[:8]}...")

        country_col = find_country_col(headers)
        year_col = find_year_col(headers)
        month_col = find_month_col(headers)

        col_mapping = {}
        for i, h in enumerate(headers):
            if i in (country_col, year_col, month_col):
                continue
            metric = map_column(h)
            if metric:
                col_mapping[i] = metric

        print(f"[sync] Country col: {country_col}, Month col: {month_col}, Year col: {year_col}")
        print(f"[sync] Mapped columns: {col_mapping}")

        if country_col is not None:
            return parse_with_country(rows[1:], headers, country_col, month_col, year_col, col_mapping)
        elif month_col is not None:
            return parse_platform_level(rows[1:], headers, month_col, year_col, col_mapping)

    return {}


def find_country_col(headers: list[str]) -> int | None:
    hints = ["country", "страна", "geo", "гео"]
    for i, h in enumerate(headers):
        low = h.lower()
        for hint in hints:
            if hint in low:
                return i
    return None


def find_month_col(headers: list[str]) -> int | None:
    hints = ["month", "месяц"]
    for i, h in enumerate(headers):
        low = h.lower()
        for hint in hints:
            if hint in low:
                return i
    return None


def find_year_col(headers: list[str]) -> int | None:
    hints = ["year", "год"]
    for i, h in enumerate(headers):
        low = h.lower()
        for hint in hints:
            if hint in low:
                return i
    return None


MONTH_ORDER = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}
MONTH_SHORT = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}
MONTH_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
    7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}


def parse_month(raw) -> int | None:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if s in MONTH_ORDER:
        return MONTH_ORDER[s]
    for name, num in MONTH_ORDER.items():
        if name.startswith(s) or s.startswith(name[:3]):
            return num
    try:
        n = int(s)
        if 1 <= n <= 12:
            return n
    except ValueError:
        pass
    return None


def parse_with_country(rows, headers, country_col, month_col, year_col, col_mapping):
    """Parse export grouped by country (and optionally by month)."""
    result = {}
    current_year = None

    for row in rows:
        cells = list(row)
        if len(cells) <= country_col:
            continue

        if year_col is not None and cells[year_col]:
            try:
                current_year = int(cells[year_col])
            except (ValueError, TypeError):
                pass

        raw_country = str(cells[country_col] or "").strip()
        if not raw_country or raw_country.lower() in ("total", "итого", "всего"):
            continue
        if raw_country.lower().startswith("примен"):
            break

        geo = resolve_geo(raw_country)
        if not geo:
            print(f"[sync] ⚠ Unknown country: '{raw_country}' — skipped")
            continue

        if month_col is not None:
            raw_month = str(cells[month_col] or "").strip()
            if raw_month.lower() in ("total", "итого", "всего", ""):
                continue

        if geo not in result:
            result[geo] = {}

        for col_idx, metric in col_mapping.items():
            if col_idx < len(cells):
                val = parse_number(cells[col_idx])
                if val is not None:
                    result[geo][metric] = val

    print(f"[sync] Parsed {len(result)} GEOs with country grouping")
    return result


def parse_platform_level(rows, headers, month_col, year_col, col_mapping):
    """Parse platform-level data (no country grouping) — extract per-month totals."""
    months_data = {}
    current_year = None

    for row in rows:
        cells = list(row)

        if year_col is not None and cells[year_col]:
            try:
                current_year = int(cells[year_col])
            except (ValueError, TypeError):
                pass

        if month_col is None or month_col >= len(cells):
            continue

        raw_month = str(cells[month_col] or "").strip()
        if raw_month.lower() in ("total", "итого", "всего", ""):
            continue
        if raw_month.lower().startswith("примен"):
            break

        month_num = parse_month(raw_month)
        if month_num is None:
            continue

        year = current_year or datetime.now().year
        key = (year, month_num)
        data = {}
        for col_idx, metric in col_mapping.items():
            if col_idx < len(cells):
                val = parse_number(cells[col_idx])
                if val is not None:
                    data[metric] = val

        if data:
            months_data[key] = data

    if not months_data:
        return {}

    sorted_months = sorted(months_data.keys())
    latest = sorted_months[-1]
    latest_data = months_data[latest]
    prev_data = months_data[sorted_months[-2]] if len(sorted_months) > 1 else {}

    print(f"[sync] Platform-level: {len(sorted_months)} months, latest={latest}")
    print(f"[sync] Latest month metrics: {list(latest_data.keys())}")

    return {
        "_platform_months": months_data,
        "_latest_month": latest,
        "_prev_month": sorted_months[-2] if len(sorted_months) > 1 else None,
        "_sorted_months": sorted_months,
    }


# ─── Merge into data.json ───

def compute_derived_metrics(raw: dict, prev_raw: dict | None = None) -> dict:
    """Compute dashboard metrics from raw WebManagement fields."""
    m = dict(raw)

    deps = raw.get("deposits_total")
    wds = raw.get("withdrawals_total")
    if deps and wds and deps > 0:
        m["cashout_pct"] = round(wds / deps * 100, 1)
    if deps and wds:
        m["ngr"] = round(deps - wds, 0)

    ggr = raw.get("ggr")
    if ggr is not None:
        m["ngr"] = round(ggr, 0)

    inout = raw.get("in_out")
    if inout is not None:
        m["ngr"] = round(inout, 0)

    reg_conv = raw.get("reg_ftd_conv")
    if reg_conv is not None and reg_conv < 1:
        m["reg_ftd_conv"] = round(reg_conv * 100, 1)

    ftd = raw.get("ftd_count")
    regs = raw.get("registrations")
    if ftd and regs and regs > 0 and "reg_ftd_conv" not in raw:
        m["reg_ftd_conv"] = round(ftd / regs * 100, 1)

    if prev_raw:
        prev_ngr = prev_raw.get("ngr") or prev_raw.get("in_out") or prev_raw.get("ggr")
        cur_ngr = m.get("ngr")
        if cur_ngr is not None and prev_ngr and prev_ngr != 0:
            m["ngr_mom"] = round((cur_ngr - prev_ngr) / abs(prev_ngr) * 100, 1)

    for k in ("deposits_total", "withdrawals_total", "in_out", "ggr", "inout_pct",
              "rtp_casino", "registrations", "cr2dep", "cr3dep", "cr4dep", "cr5dep"):
        m.pop(k, None)

    return m


def merge_into_dataset(existing: dict, parsed: dict) -> dict:
    ds = json.loads(json.dumps(existing))

    if "_platform_months" in parsed:
        pm = parsed["_platform_months"]
        sorted_months = parsed["_sorted_months"]
        latest_key = parsed["_latest_month"]
        prev_key = parsed["_prev_month"]

        latest_data = pm[latest_key]
        prev_data = pm.get(prev_key, {}) if prev_key else {}
        derived = compute_derived_metrics(latest_data, prev_data)

        ngr_series = []
        labels = []
        for ym in sorted_months:
            md = pm[ym]
            ngr = md.get("in_out") or md.get("ggr") or md.get("ngr", 0)
            ngr_series.append(round(ngr))
            labels.append(MONTH_SHORT.get(ym[1], str(ym[1])))

        ds["plat_trend"] = ngr_series
        ds["trend_labels"] = labels
        latest_y, latest_m = latest_key
        ds["meta"]["month"] = f"{MONTH_RU.get(latest_m, '')} {latest_y}"
        ds["meta"]["month_short"] = MONTH_SHORT.get(latest_m, "")
        if prev_key:
            ds["meta"]["prev_month_short"] = MONTH_SHORT.get(prev_key[1], "").lower()[:3]

        print(f"[sync] Updated platform trend: {len(ngr_series)} months")
    else:
        if "data" not in ds:
            ds["data"] = {}

        for geo, metrics in parsed.items():
            if geo.startswith("_"):
                continue
            derived = compute_derived_metrics(metrics)
            if geo not in ds["data"]:
                ds["data"][geo] = {}
            for k, v in derived.items():
                ds["data"][geo][k] = v
            print(f"[sync] ✓ {geo}: {len(derived)} metrics")

    ds["meta"]["updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return ds


# ─── Main ───

def main():
    if not DRIVE_FOLDER_ID:
        print("[sync] ERROR: DRIVE_FOLDER_ID not set")
        sys.exit(1)

    files = list_drive_files(DRIVE_FOLDER_ID)
    if not files:
        print("[sync] ERROR: No files found in Drive folder")
        sys.exit(1)

    target = pick_best_file(files)
    if not target:
        print("[sync] ERROR: No .xlsx or .csv files found")
        sys.exit(1)

    print(f"[sync] Selected: {target['name']}")

    ext = os.path.splitext(target["name"])[1].lower()
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp_path = tmp.name

    try:
        download_drive_file(target["id"], tmp_path)

        if ext in (".xlsx", ".xls"):
            parsed = parse_xlsx(tmp_path)
        elif ext in (".csv", ".tsv"):
            with open(tmp_path, "r", encoding="utf-8-sig") as f:
                from sync_google_sheet import parse_sheet
                parsed_data = parse_sheet(f.read())
                parsed = parsed_data
        else:
            print(f"[sync] ERROR: Unsupported format: {ext}")
            sys.exit(1)
    finally:
        os.unlink(tmp_path)

    if not parsed:
        print("[sync] WARNING: No data parsed")
        sys.exit(1)

    if os.path.exists(DATA_JSON_PATH):
        with open(DATA_JSON_PATH, "r", encoding="utf-8") as f:
            existing = json.load(f)
    else:
        existing = {
            "meta": {"month": "", "month_short": "", "prev_month_short": "", "version": "v29", "updated": ""},
            "flags": {}, "trend_labels": [], "plat_trend": [], "prev": {}, "trend": {}, "data": {},
        }

    merged = merge_into_dataset(existing, parsed)

    with open(DATA_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    print(f"[sync] ✅ Written to {DATA_JSON_PATH}")


if __name__ == "__main__":
    main()
